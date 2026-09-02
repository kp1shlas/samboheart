"""
Сервис для импорта детей из Excel-файлов
"""
import openpyxl
from datetime import datetime
from django.contrib.auth.models import User
from django.utils.crypto import get_random_string
from core.models import Child, ParentProfile, Group, ChildEnrollment


def generate_username(full_name):
    """Генерирует логин на основе ФИО"""
    # Транслитерация
    translit = {
        'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e',
        'ё': 'yo', 'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k',
        'л': 'l', 'м': 'm', 'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r',
        'с': 's', 'т': 't', 'у': 'u', 'ф': 'f', 'х': 'h', 'ц': 'c',
        'ч': 'ch', 'ш': 'sh', 'щ': 'sch', 'ъ': '', 'ы': 'y', 'ь': '',
        'э': 'e', 'ю': 'yu', 'я': 'ya', ' ': '_',
    }
    
    username = ''.join(translit.get(c.lower(), c) for c in full_name)
    username = username.strip('_')
    
    # Добавляем случайные цифры если username занят
    base_username = username
    counter = 1
    while User.objects.filter(username=username).exists():
        username = f'{base_username}_{counter}'
        counter += 1
    
    return username


def parse_date(date_value):
    """Парсит дату из разных форматов"""
    if not date_value:
        return None
    
    # Если это уже datetime
    if isinstance(date_value, datetime):
        return date_value.date()
    
    # Если строка
    if isinstance(date_value, str):
        date_value = date_value.strip()
        
        # Пробуем разные форматы
        formats = [
            '%d.%m.%Y',  # 15.03.2015
            '%d-%m-%Y',  # 15-03-2015
            '%Y-%m-%d',  # 2015-03-15
            '%d/%m/%Y',  # 15/03/2015
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_value, fmt).date()
            except ValueError:
                continue
    
    return None


def import_children_from_excel(file_obj):
    """
    Импортирует детей из Excel-файла
    
    Ожидаемые колонки:
    A: ФИО (обязательно)
    B: Дата рождения (ДД.ММ.ГГГГ)
    C: Логин (опционально)
    D: Пароль (опционально)
    E: Группа (опционально)
    F: Родитель (username, опционально)
    
    Returns:
        dict: {
            'success': int,
            'errors': list,
            'skipped': int,
        }
    """
    result = {
        'success': 0,
        'errors': [],
        'skipped': 0,
        'children': [],
    }
    
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active
        
        # Проверяем наличие данных
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        if not rows:
            result['errors'].append('Файл пустой или содержит только заголовки')
            return result
        
        for row_num, row in enumerate(rows, start=2):
            try:
                # Извлекаем данные
                full_name = str(row[0] or '').strip() if row[0] else ''
                birth_date_raw = row[1] if len(row) > 1 else None
                username = str(row[2] or '').strip() if len(row) > 2 and row[2] else ''
                password = str(row[3] or '').strip() if len(row) > 3 and row[3] else ''
                group_name = str(row[4] or '').strip() if len(row) > 4 and row[4] else ''
                parent_username = str(row[5] or '').strip() if len(row) > 5 and row[5] else ''
                
                # Проверка обязательных полей
                if not full_name:
                    result['errors'].append(f'Строка {row_num}: ФИО не указано')
                    result['skipped'] += 1
                    continue
                
                # Парсим дату рождения
                birth_date = parse_date(birth_date_raw)
                
                # Генерируем логин если не указан
                if not username:
                    username = generate_username(full_name)
                
                # Генерируем пароль если не указан
                if not password:
                    password = get_random_string(8)
                
                # Проверяем, существует ли пользователь
                if User.objects.filter(username=username).exists():
                    result['errors'].append(
                        f'Строка {row_num}: Логин "{username}" уже существует (ФИО: {full_name})'
                    )
                    result['skipped'] += 1
                    continue
                
                # Создаём пользователя
                first_name = full_name.split()[1] if len(full_name.split()) > 1 else ''
                last_name = full_name.split()[0] if full_name.split() else ''
                
                user = User.objects.create_user(
                    username=username,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                )
                
                # Ищем родителя если указан
                parent = None
                if parent_username:
                    try:
                        parent_user = User.objects.get(username=parent_username)
                        parent, _ = ParentProfile.objects.get_or_create(user=parent_user)
                    except User.DoesNotExist:
                        result['errors'].append(
                            f'Строка {row_num}: Родитель "{parent_username}" не найден (ФИО: {full_name})'
                        )
                
                # Создаём ребёнка
                child = Child.objects.create(
                    user=user,
                    parent=parent,
                    full_name=full_name,
                    birth_date=birth_date,
                    is_active=True,
                )
                
                # Записываем в группу если указана
                if group_name:
                    try:
                        group = Group.objects.get(name__iexact=group_name)
                        ChildEnrollment.objects.get_or_create(
                            child=child,
                            group=group,
                            defaults={'is_active': True, 'remaining_lessons': 0}
                        )
                    except Group.DoesNotExist:
                        result['errors'].append(
                            f'Строка {row_num}: Группа "{group_name}" не найдена (ФИО: {full_name})'
                        )
                    except Group.MultipleObjectsReturned:
                        result['errors'].append(
                            f'Строка {row_num}: Найдено несколько групп с именем "{group_name}" (ФИО: {full_name})'
                        )
                
                result['success'] += 1
                result['children'].append({
                    'full_name': full_name,
                    'username': username,
                    'password': password,
                })
                
            except Exception as e:
                result['errors'].append(f'Строка {row_num}: Ошибка - {str(e)}')
                result['skipped'] += 1
                continue
        
    except Exception as e:
        result['errors'].append(f'Ошибка чтения файла: {str(e)}')
    
    return result


def create_excel_template():
    """
    Создаёт шаблон Excel-файла для скачивания
    
    Returns:
        BytesIO: файл шаблона
    """
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Дети'
    
    # Заголовки
    headers = [
        'ФИО *',
        'Дата рождения (ДД.ММ.ГГГГ)',
        'Логин (авто если пусто)',
        'Пароль (авто если пусто)',
        'Группа',
        'Родитель (username)',
    ]
    ws.append(headers)
    
    # Примеры
    examples = [
        ['Иванов Иван Иванович', '15.03.2015', 'ivanov_ii', '123456', 'Самбо 7-9 лет', 'ivanov_parent'],
        ['Петрова Анна Сергеевна', '20.05.2016', '', '', 'Самбо 7-9 лет', ''],
        ['Сидоров Пётр Петрович', '10.09.2014', '', '', '', ''],
    ]
    
    for example in examples:
        ws.append(example)
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25
    
    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output